"""Orchestrator for the tk-vn product sheet pipeline.

  prepare  <xlsx> <work.json>
      Run deterministic transforms (brand, stock, video, sku, url normalization)
      directly on a fresh copy, then emit a work.json of *agent* tasks
      (translations to write, images to visually inspect). Does NOT call the API.

  finalize <xlsx> <work.json> <out.xlsx>
      Take the agent-filled results in work.json and write them into the sheet
      (Vietnamese title/variants, replaced image URLs, weight, dims, rewritten
      description). Backs up the source when out == source.

work.json shape:
  {
    "date": "YYYYMMDD",                 # sku prefix
    "start_seq": 1,                     # sku sequence start
    "rows": [
      {
        "row_index": 2,
        "sku": "2026070100001",
        "translate": {"B": <vi title>, "G": <vi>, "H": <vi>, "I": <vi>,
                      "J": <vi>, "K": <vi>, "L": <vi>},   # agent fills
        "images": [
          {"source": "main"|"sub"|"variant"|"desc",
           "col": "R"|"S"..|"AC"|"C", "orig": <url>,
           "decision": <"keep"|"delete"|"regen">,  # agent fills
           "new_url": <url or "">,                  # agent fills when regen
           "vi_text": <text or "">,                 # agent fills when regen w/ text
           "weight_kg": <num|None>, "l":<num|None>, "w":<num|None>, "h":<num|None>}
        ]
      }
    ]
  }
"""
from __future__ import annotations

import json
import shutil
import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

import sheet_io  # noqa: E402  (same dir)

SHEET = sheet_io.SHEET_NAME
COL = sheet_io.COL
SUB_COLS = COL["sub_imgs"]  # S..Z


def _today_str() -> str:
    return date.today().strftime("%Y%m%d")


def prepare(xlsx_path: str, work_json: str) -> None:
    """Read-only: emit work.json of agent tasks. Does NOT modify the xlsx.

    Deterministic transforms (brand/stock/video/sku/url-normalize) are applied
    later in `finalize` so there is a single write step.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active

    today = _today_str()
    seq = 1
    rows = []
    for r in range(2, ws.max_row + 1):
        if all((ws.cell(row=r, column=c).value in (None, "")) for c in range(1, ws.max_column + 1)):
            continue
        sku = f"{today}{seq:05d}"
        seq += 1

        def read_url(letter: str) -> str:
            v = ws.cell(row=r, column=sheet_io.col_idx(letter)).value
            return sheet_io.normalize_url(v)

        main = read_url(COL["main_img"])
        subs = [read_url(c) for c in SUB_COLS]
        variant = read_url(COL["variant_img"])
        desc_html = ws.cell(row=r, column=sheet_io.col_idx(COL["desc"])).value
        desc_urls = sheet_io.extract_img_urls(desc_html)

        images = []
        if main:
            images.append({"source": "main", "col": COL["main_img"], "orig": main,
                           "decision": "", "new_url": "", "vi_text": "",
                           "weight_kg": None, "l": None, "w": None, "h": None})
        for letter, url in zip(SUB_COLS, subs):
            if url:
                images.append({"source": "sub", "col": letter, "orig": url,
                               "decision": "", "new_url": "", "vi_text": "",
                               "weight_kg": None, "l": None, "w": None, "h": None})
        if variant:
            images.append({"source": "variant", "col": COL["variant_img"], "orig": variant,
                           "decision": "", "new_url": "", "vi_text": "",
                           "weight_kg": None, "l": None, "w": None, "h": None})
        for url in desc_urls:
            images.append({"source": "desc", "col": COL["desc"], "orig": url,
                           "decision": "", "new_url": "", "vi_text": "",
                           "weight_kg": None, "l": None, "w": None, "h": None})

        # 提取C列文字部分(去掉img标签后的文字HTML),供agent翻译去品牌
        desc_text = sheet_io.extract_text_content(desc_html)

        rows.append({
            "row_index": r,
            "sku": sku,
            "translate": {"B": "", "G": "", "H": "", "I": "", "J": "", "K": "", "L": ""},
            "desc_text_original": desc_text,   # 原始C列文字(含中文/品牌名), 只读参考
            "desc_text_vi": "",                # agent填: 去品牌+翻译越南语后的文字HTML
            "images": images,
        })

    work = {"date": today, "start_seq": 1, "rows": rows}
    Path(work_json).write_text(json.dumps(work, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"prepare: {len(rows)} rows -> {work_json}")
    print("Fill translate/images in work.json, then run `finalize` to write the sheet.")


def finalize(xlsx_path: str, work_json: str, out_xlsx: str) -> None:
    work = json.loads(Path(work_json).read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active

    img_cols = [COL["main_img"]] + SUB_COLS + [COL["variant_img"]]
    for row in work["rows"]:
        r = row["row_index"]

        # --- deterministic transforms (applied here so there is one write step) ---
        ws.cell(row=r, column=sheet_io.col_idx(COL["brand"])).value = None  # 品牌留空
        ws.cell(row=r, column=sheet_io.col_idx(COL["stock"])).value = 30
        ws.cell(row=r, column=sheet_io.col_idx(COL["video"])).value = None  # delete video link
        ws.cell(row=r, column=sheet_io.col_idx(COL["sku"])).value = row["sku"]
        # normalize standalone image-cell URLs as the baseline
        for letter in img_cols:
            ci = sheet_io.col_idx(letter)
            v = ws.cell(row=r, column=ci).value
            nv = sheet_io.normalize_url(v)
            ws.cell(row=r, column=ci).value = nv or None

        # --- translations ---
        for letter, val in row.get("translate", {}).items():
            if val:
                ws.cell(row=r, column=sheet_io.col_idx(letter)).value = val

        # --- image decisions ---
        desc_delete: set[str] = set()
        desc_replace: dict[str, str] = {}
        collected_weight = None
        collected_l = collected_w = collected_h = None
        for img in row["images"]:
            dec = (img.get("decision") or "").lower()
            orig = img["orig"]
            new_url = img.get("new_url") or ""
            if img.get("weight_kg") is not None:
                collected_weight = img["weight_kg"]
            if img.get("l") is not None:
                collected_l = img["l"]
            if img.get("w") is not None:
                collected_w = img["w"]
            if img.get("h") is not None:
                collected_h = img["h"]

            if dec == "keep":
                continue
            if dec == "delete":
                if img["col"] == COL["desc"]:
                    desc_delete.add(orig)
                else:
                    ws.cell(row=r, column=sheet_io.col_idx(img["col"])).value = None
                continue
            if dec == "regen" and new_url:
                if img["col"] == COL["desc"]:
                    desc_replace[orig] = new_url
                else:
                    ws.cell(row=r, column=sheet_io.col_idx(img["col"])).value = new_url
                continue

        # rewrite description (C): 文字部分用翻译后的, 图片部分删无关+换清洗后, 保持顺序
        desc_ci = sheet_io.col_idx(COL["desc"])
        cur_html = ws.cell(row=r, column=desc_ci).value
        cur_urls = sheet_io.extract_img_urls(cur_html)  # normalizes too
        new_urls = [desc_replace.get(u, u) for u in cur_urls if u not in desc_delete]
        # 文字: agent填了越南语翻译就用翻译, 否则保留原文字
        vi_text = row.get("desc_text_vi") or ""
        if not vi_text:
            vi_text = sheet_io.extract_text_content(cur_html)
        ws.cell(row=r, column=desc_ci).value = sheet_io.build_description_all(vi_text, new_urls)

        # physical attrs
        if collected_weight is not None:
            ws.cell(row=r, column=sheet_io.col_idx(COL["weight"])).value = round(float(collected_weight), 3)
        if collected_l is not None:
            ws.cell(row=r, column=sheet_io.col_idx(COL["length"])).value = collected_l
        if collected_w is not None:
            ws.cell(row=r, column=sheet_io.col_idx(COL["width"])).value = collected_w
        if collected_h is not None:
            ws.cell(row=r, column=sheet_io.col_idx(COL["height"])).value = collected_h

    # --- column cleanup: 删「本地展示价」空列 + 「价格(站点币种)」改名「本地展示价」 ---
    # 按标题定位,不硬编码列号(先删空列可能改变列号,所以先找再删)
    header_row = 1
    price_col = None
    local_price_col = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=header_row, column=c).value or "").strip()
        if h.startswith("价格") or "价格(站点币种)" in h or "价格（站点币种）" in h:
            price_col = c
        if h == "本地展示价" or h.startswith("本地展示价"):
            local_price_col = c
    # 删除独立的「本地展示价」空列(仅当它与价格列不是同一列时)
    if local_price_col and local_price_col != price_col:
        ws.delete_cols(local_price_col, 1)
        # 删列后,若价格列在被删列之后,列号左移1
        if price_col and price_col > local_price_col:
            price_col -= 1
    # 「价格(站点币种)」标题改名为「本地展示价」
    if price_col:
        ws.cell(row=header_row, column=price_col).value = "本地展示价"

    if Path(out_xlsx).resolve() == Path(xlsx_path).resolve():
        bak = Path(xlsx_path).with_suffix(Path(xlsx_path).suffix + ".bak")
        shutil.copy2(xlsx_path, bak)
        print(f"backed up original -> {bak}")
    wb.save(out_xlsx)
    print(f"finalize: wrote results -> {out_xlsx}")


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(__doc__)
        return 2
    if argv[1] == "prepare":
        prepare(argv[2], argv[3])
    elif argv[1] == "finalize":
        finalize(argv[2], argv[3], argv[4])
    else:
        print(f"unknown command: {argv[1]}")
        return 2
    return 0


if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    sys.exit(main(sys.argv))
