"""Parallel batch image processor — fast mode (agent-cluster-like concurrency).

Usage:
  python scripts/batch_process.py "<xlsx>" [--doubao-key KEY] [--hfsy-key KEY] [--agnes-key KEY]
  Optional: --audit-workers 24  --gen-workers 24  --gen-size 4K

一条命令全自动 (zero glue code needed):
  1. prepare (deterministic: brand留空/stock=30/SKU/video清空/URL归一)
  2. auto-translate (标题/变种/描述文字 → 越南语去品牌, minimax-m3, 并发)
  3. Vision audit all unique images (N parallel) → structured JSON per image
  4. Batch image gen for brand/logo/watermark/text only
     (M parallel, edits → nano-banana-2 → Doubao → generations fallback chain)
  5. Share main/sub URLs to all variant rows
  6. finalize (write all results + 45→35列对齐)

Concurrency = Python ThreadPoolExecutor (no agent runtime needed, open-source
friendly). Bottleneck is API wait; 429/5xx auto-backoff-retry (自适应限流).
"""
from __future__ import annotations

import argparse
import base64
import hashlib
import json
import os
import random
import re
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any

import requests

sys.path.insert(0, str(Path(__file__).resolve().parent))
import run_pipeline  # noqa: E402
from run_pipeline import prepare as rp_prepare, finalize as rp_finalize

PROMPT_FILE = Path(__file__).resolve().parent.parent / '图片生成提示词.md'

def _load_prompt() -> str:
    """Read the shared image-cleaning prompt from the project root markdown file."""
    text = PROMPT_FILE.read_text(encoding='utf-8')
    # The first line is '# 提示词'; everything after is the prompt body.
    _, _, body = text.partition('
')
    return body.lstrip('
')

PROMPT = _load_prompt()

AUDIT_SYSTEM = (
    "Audit product image. Output JSON only: "
    '{"has_brand_name":bool,"brand_names_found":[],"has_logo":bool,'
    '"has_watermark":bool,"has_chinese_text":bool,"is_promo_banner":bool,'
    '"needs_cleaning":bool,"cleaning_reason":"brand|logo|watermark|text|promo|none"}'
)

# 批量翻译的 system prompt (标题/变种/描述文字一次性翻译, 去品牌+越南语)
TRANSLATE_SYSTEM = (
    "你是TikTok越南站跨境电商翻译专家。把给定JSON里的中文字段翻译成越南语,规则:\n"
    "1. title(产品标题): 品类名词开头, ≤80字符, 删除'原装/原厂/正品/专柜/官方', "
    "品牌词改成'phù hợp với [品牌]'形式, 避免侵权.\n"
    "2. vname/vval(变种属性名/值): 删品牌名+译越南语. 颜色分类→Phân loại màu, 商品规格→Quy cách sản phẩm.\n"
    "3. desc_text(描述文字HTML): 删品牌名(如具体品牌改成通用词), 译越南语, 保留<p>等HTML标签.\n"
    "只输出JSON, 键与输入完全一致, 值为越南语译文. 空值保持空."
)


def translate_batch(items: dict, ark_key: str, timeout: int = 120, max_retries: int = 4) -> dict:
    """批量翻译一组中文字段→越南语. items是 {key: 中文} 的dict, 返回 {key: 越南语}.
    用 minimax-m3 文本模型. 429/5xx 指数退避重试. 失败返回空dict(agent可后续补)."""
    if not items or not ark_key:
        return {}
    import time as _t
    for attempt in range(max_retries):
        try:
            resp = requests.post(
                "https://ark.cn-beijing.volces.com/api/coding/v1/chat/completions",
                headers={"Authorization": f"Bearer {ark_key}"},
                json={
                    "model": "minimax-m3",
                    "messages": [
                        {"role": "system", "content": TRANSLATE_SYSTEM},
                        {"role": "user", "content": json.dumps(items, ensure_ascii=False)},
                    ],
                    "max_tokens": 4000,
                },
                timeout=timeout,
            )
            if resp.status_code == 429 or resp.status_code >= 500:
                _t.sleep(min(2 ** attempt * 3, 30))
                continue
            content = resp.json()["choices"][0]["message"]["content"]
            m = re.search(r"\{.*\}", content, re.DOTALL)
            if m:
                return json.loads(m.group(0))
            return {}
        except Exception:
            _t.sleep(min(2 ** attempt * 2, 20))
    return {}


# ── Helpers: key rotation, checkpoint, progress ────────────────────────────


class KeyRoundRobin:
    """Thread-safe round-robin key distributor."""

    def __init__(self, keys: list[str]):
        self._keys = keys
        self._counter = 0
        self._lock = threading.Lock()

    def next_key(self) -> str | None:
        if not self._keys:
            return None
        with self._lock:
            key = self._keys[self._counter % len(self._keys)]
            self._counter += 1
            return key


def checkpoint_save(path: str, completed: dict, total: int) -> None:
    """Save current gen progress to checkpoint JSON."""
    data = {
        "checkpoint_version": 1,
        "completed": completed,
        "total": total,
        "timestamp": time.time(),
    }
    Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def checkpoint_load(path: str) -> dict | None:
    """Load checkpoint. Returns None if not found or corrupted."""
    p = Path(path)
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None


def _print_progress(done: int, total: int, elapsed: float, label: str = "") -> None:
    """Print a text progress bar to stderr."""
    if total == 0:
        pct = 100.0
    else:
        pct = done / total * 100
    bar_len = 30
    filled = int(bar_len * done / max(total, 1))
    bar = "=" * filled + "-" * (bar_len - filled)
    eta_secs = 0
    if done > 0 and elapsed > 0:
        eta_secs = (elapsed / done) * (total - done)
    if eta_secs > 60:
        eta_str = f"{int(eta_secs // 60)}m{int(eta_secs % 60)}s"
    else:
        eta_str = f"{int(eta_secs)}s"
    prefix = f"  [{label}] " if label else "  "
    sys.stderr.write(f"\r{prefix}[{bar}] {done}/{total} ({pct:.0f}%) ETA {eta_str}  ")
    sys.stderr.flush()
    if done >= total:
        sys.stderr.write("\n")
        sys.stderr.flush()


def _try_key_chain(callable_fn, max_attempts: int = 3) -> Any | None:
    """Try callable_fn() with exponential backoff on retryable errors."""
    last_err = None
    for attempt in range(max_attempts):
        try:
            result = callable_fn()
            if result is not None:
                return result
        except requests.exceptions.HTTPError as e:
            if e.response is not None and e.response.status_code in (429, 500, 502, 503, 504):
                delay = 1.0 * (2 ** attempt) * (0.5 + random.random())
                time.sleep(delay)
                last_err = e
                continue
            return None  # non-retryable (401, 400, etc.)
        except Exception as e:
            delay = 1.0 * (2 ** attempt) * (0.5 + random.random())
            time.sleep(delay)
            last_err = e
    return None


# ── API call helpers ──────────────────────────────────────


def vision_audit(url: str, agnes_key: str, timeout: int = 60,
                 ark_key: str = "") -> dict:
    """Audit a single image by URL. No local download — sends URL directly.
    Primary: minimax-m3 via Volcengine coding endpoint (better recall on
    brand/logo/watermark). Fallback: Agnes 2.0 flash. Returns structured dict.
    """
    # --- Primary: minimax-m3 (Volcengine) ---
    if ark_key:
        try:
            resp = requests.post(
                "https://ark.cn-beijing.volces.com/api/coding/v1/chat/completions",
                headers={"Authorization": f"Bearer {ark_key}"},
                json={
                    "model": "minimax-m3",
                    "messages": [{"role": "user", "content": [
                        {"type": "text", "text": AUDIT_SYSTEM + "\nBe thorough: check every corner for small logos, semi-transparent watermarks, tiny brand marks."},
                        {"type": "image_url", "image_url": {"url": url}}
                    ]}],
                    "max_tokens": 500,
                },
                timeout=timeout,
            )
            content = resp.json()["choices"][0]["message"]["content"]
            m = re.search(r"\{.*\}", content, re.DOTALL)
            if m:
                d = json.loads(m.group(0))
                if "needs_cleaning" in d or "has_brand_name" in d:
                    return d
        except Exception:
            pass
    # --- Fallback: Agnes 2.0 flash ---
    try:
        resp = requests.post(
            "https://apihub.agnes-ai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {agnes_key}"},
            json={
                "model": "agnes-2.0-flash",
                "messages": [
                    {"role": "system", "content": AUDIT_SYSTEM},
                    {"role": "user", "content": [
                        {"type": "text", "text": "Audit this image for brand names, logos, watermarks, or Chinese text. Be thorough."},
                        {"type": "image_url", "image_url": {"url": url}}
                    ]}
                ],
                "max_tokens": 400,
            },
            timeout=timeout,
        )
        content = resp.json()["choices"][0]["message"]["content"]
        m = re.search(r"\{.*\}", content, re.DOTALL)
        if m:
            return json.loads(m.group(0))
    except Exception:
        pass
    # On total failure, conservatively assume needs cleaning (safer for IP)
    return {"needs_cleaning": True, "cleaning_reason": "unknown", "is_promo_banner": False}


def _to_data_uri(url: str) -> str | None:
    """Download image and return base64 data URI (for gw.alicdn that blocks API fetch)."""
    try:
        r = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code == 200:
            return f"data:image/jpeg;base64,{base64.b64encode(r.content).decode()}"
    except Exception:
        pass
    return None


def nano_gen(url: str, key: str, size: str = "4K", timeout: int = 300) -> str | None:
    """Primary: nano-banana-2 (up to 4K). Returns URL or None."""
    try:
        img_uri = url
        if "gw.alicdn" in url:
            img_uri = _to_data_uri(url) or url
        resp = requests.post(
            "https://www.hfsyapi.cn/v1beta/models/nano-banana-2:generateContent",
            headers={"Authorization": f"Bearer {key}", "User-Agent": "curl/7.68.0"},
            json={
                "contents": [{"parts": [
                    {"text": PROMPT},
                    {"fileData": {"mimeType": "image/jpeg", "fileUri": img_uri}}
                ]}],
                "generationConfig": {"imageConfig": {"imageSize": size, "aspectRatio": "1:1"}},
            },
            timeout=timeout,
        )
        data = resp.json()
        for cand in data.get("candidates", []):
            for p in cand.get("content", {}).get("parts", []):
                fd = p.get("fileData")
                if fd and fd.get("fileUri"):
                    return fd["fileUri"]
    except Exception:
        pass
    return None


def doubao_gen(url: str, key: str, size: str = "2K", timeout: int = 180) -> str | None:
    """Fallback 1: Doubao Seedream 5.0. Returns URL or None."""
    try:
        resp = requests.post(
            "https://ark.cn-beijing.volces.com/api/v3/images/generations",
            headers={"Authorization": f"Bearer {key}"},
            json={
                "model": "doubao-seedream-5-0-260128",
                "prompt": PROMPT,
                "image": url,
                "response_format": "url",
                "size": size,
                "watermark": False,
            },
            timeout=timeout,
        )
        return resp.json()["data"][0]["url"]
    except Exception:
        return None


def hfsyapi_gen(url: str, key: str, timeout: int = 300) -> str | None:
    """Fallback 2: GPT-Image-2 via hfsyapi (1K). Returns URL or None."""
    try:
        ref = url
        if "gw.alicdn" in url:
            ref = _to_data_uri(url) or url
        resp = requests.post(
            "https://www.hfsyapi.cn/v1/images/generations",
            headers={"Authorization": f"Bearer {key}", "User-Agent": "curl/7.68.0"},
            json={
                "model": "gpt-image-2",
                "prompt": PROMPT,
                "reference_images": [ref],
                "size": "1024x1024",
                "n": 1,
                "response_format": "url",
            },
            timeout=timeout,
        )
        return resp.json()["data"][0]["url"]
    except Exception:
        return None


def edit_gen(url: str, key: str, timeout: int = 300, max_retries: int = 4) -> str | None:
    """Primary: gpt-image-2 /v1/images/edits (multipart, best product preservation).
    Edit mode keeps original closer than full regeneration. Returns URL or None.
    429/5xx → exponential backoff retry (自适应限流)."""
    import io
    import time as _t
    import uuid
    try:
        img = requests.get(url if url.startswith("http") else "https:" + url,
                           timeout=60, headers={"User-Agent": "Mozilla/5.0"}).content
    except Exception:
        return None
    for attempt in range(max_retries):
        try:
            boundary = uuid.uuid4().hex
            buf = io.BytesIO()
            for n, v in {"model": "gpt-image-2", "prompt": PROMPT,
                         "size": "1024x1024", "response_format": "url"}.items():
                buf.write(f'--{boundary}\r\nContent-Disposition: form-data; name="{n}"\r\n\r\n{v}\r\n'.encode())
            buf.write(f'--{boundary}\r\nContent-Disposition: form-data; name="image"; '
                      f'filename="i.jpg"\r\nContent-Type: image/jpeg\r\n\r\n'.encode() + img + b"\r\n")
            buf.write(f"--{boundary}--\r\n".encode())
            resp = requests.post(
                "https://www.hfsyapi.cn/v1/images/edits",
                headers={"Authorization": f"Bearer {key}",
                         "Content-Type": f"multipart/form-data; boundary={boundary}",
                         "User-Agent": "curl/7.68.0"},
                data=buf.getvalue(), timeout=timeout,
            )
            if resp.status_code == 429 or resp.status_code >= 500:
                # 限流/服务端错误 → 指数退避重试
                _t.sleep(min(2 ** attempt * 3, 30))
                continue
            return resp.json()["data"][0]["url"]
        except Exception:
            _t.sleep(min(2 ** attempt * 2, 20))
    return None


# ── Main pipeline ─────────────────────────────────────────


def auto_process(xlsx_path: str, ark_key: str, hfsy_key: str, agnes_key: str,
                 work_path: str | None = None, audit_workers: int = 24,
                 gen_workers: int = 24, gen_size: str = "4K",
                 _ark_keys: list[str] | None = None,
                 _hfsy_keys: list[str] | None = None,
                 _checkpoint_path: str | None = None) -> dict[str, Any]:
    xlsx = Path(xlsx_path).resolve()
    work = Path(work_path or xlsx.with_name("work_auto.json")).resolve()

    # Multi-key setup (backward compat: single key falls through)
    ark_keys = _ark_keys or ([ark_key] if ark_key else [])
    hfsy_keys = _hfsy_keys or ([hfsy_key] if hfsy_key else [])
    agnes_keys = [agnes_key] if agnes_key else []

    xlsx = Path(xlsx_path).resolve()
    work = Path(work_path or xlsx.with_name("work_auto.json")).resolve()

    # Step 1: prepare (fast, <1s)
    print("[1/5] Prepare (deterministic transforms)...", flush=True)
    rp_prepare(str(xlsx), str(work))
    w = json.loads(work.read_text(encoding="utf-8"))

    # Step 1b: 自动翻译 (标题/变种/描述文字 → 越南语, 去品牌). 一次一行, 并发.
    print(f"[1b] Auto-translate titles/variants/desc ({audit_workers} parallel)...", flush=True)
    import openpyxl as _oxl
    import run_pipeline as _rp
    _wb = _oxl.load_workbook(str(xlsx), data_only=True)
    _ws = _wb[_rp.SHEET] if _rp.SHEET in _wb.sheetnames else _wb.active
    _CI = _rp.sheet_io.col_idx

    def translate_row(row: dict, key: str) -> dict:
        r = row["row_index"]
        src = {}
        for key_name, col in [("title", "B"), ("vname1", "G"), ("vval1", "H"),
                              ("vname2", "I"), ("vval2", "J"), ("vname3", "K"), ("vval3", "L")]:
            v = _ws.cell(row=r, column=_CI(col)).value
            if v and str(v).strip():
                src[key_name] = str(v)
        if row.get("desc_text_original"):
            src["desc_text"] = row["desc_text_original"]
        if not src:
            return {"row_index": r, "translate": {}, "desc_text_vi": ""}
        vi = translate_batch(src, key)
        tr = {}
        for key_name, col in [("title", "B"), ("vname1", "G"), ("vval1", "H"),
                              ("vname2", "I"), ("vval2", "J"), ("vname3", "K"), ("vval3", "L")]:
            if vi.get(key_name):
                tr[col] = vi[key_name]
        return {"row_index": r, "translate": tr, "desc_text_vi": vi.get("desc_text", "")}

    # Multi-key translation via round-robin
    if ark_keys:
        _rr_trans = KeyRoundRobin(ark_keys)
        _trans_lock = threading.Lock()
        trans: dict[int, dict] = {}

        def _translate_row_rr(row: dict) -> dict:
            key = _rr_trans.next_key() or ark_key
            result = translate_row(row, key)
            with _trans_lock:
                trans[result["row_index"]] = result
            return result

        with ThreadPoolExecutor(max_workers=audit_workers) as ex:
            list(ex.map(_translate_row_rr, w["rows"]))
        for row in w["rows"]:
            t = trans.get(row["row_index"])
            if t:
                row["translate"] = t["translate"]
                if t["desc_text_vi"]:
                    row["desc_text_vi"] = t["desc_text_vi"]
        n_tr = sum(1 for t in trans.values() if t["translate"])
        print(f"   translated {n_tr} rows", flush=True)
    else:
        print("   跳过(无ark_key), 翻译需agent手动填work.json", flush=True)

    # Collect unique image URLs (dedup)
    unique_urls: dict[str, dict] = {}
    for row in w["rows"]:
        for img in row["images"]:
            url = img["orig"]
            if url and url not in unique_urls:
                unique_urls[url] = img
    all_urls = list(unique_urls.keys())
    print(f"   {len(all_urls)} unique images across {len(w['rows'])} rows", flush=True)

    # Step 2: Vision audit (parallel — cluster-like)
    print(f"[2/5] Vision audit ({audit_workers} parallel)...", flush=True)
    audits: dict[str, dict] = {}
    with ThreadPoolExecutor(max_workers=audit_workers) as ex:
        fut = {ex.submit(vision_audit, url, agnes_key, 60, ark_key): url for url in all_urls}
        for f in as_completed(fut):
            audits[fut[f]] = f.result()

    def classify(a: dict, source: str) -> str:
        """Classify an image. `delete` ONLY applies to desc(C) column images."""
        needs_clean = (a.get("needs_cleaning") or a.get("has_brand_name")
                       or a.get("has_logo") or a.get("has_watermark")
                       or a.get("has_chinese_text"))
        if source == "desc":
            if a.get("is_promo_banner"):
                return "delete"
            return "regen" if needs_clean else "keep"
        else:
            return "regen" if needs_clean else "keep"

    url_source: dict[str, str] = {}
    for row in w["rows"]:
        for img in row["images"]:
            u, s = img["orig"], img.get("source", "desc")
            if u not in url_source or url_source[u] == "desc":
                url_source[u] = s

    decisions = {url: classify(audits.get(url, {}), url_source.get(url, "desc"))
                 for url in all_urls}
    n_del = sum(1 for d in decisions.values() if d == "delete")
    n_regen = sum(1 for d in decisions.values() if d == "regen")
    n_keep = sum(1 for d in decisions.values() if d == "keep")
    print(f"   delete(desc only)={n_del} regen={n_regen} keep={n_keep}", flush=True)

    # Step 3: Generate only for regen (parallel, fallback chain)
    to_gen = [u for u in all_urls if decisions.get(u) == "regen"]
    print(f"[3/5] Image gen for {len(to_gen)} images ({gen_workers} parallel)...", flush=True)

    # Checkpoint/resume
    ckpt_path = _checkpoint_path or str(work.with_name("_checkpoint.json"))
    cp = checkpoint_load(ckpt_path)
    cp_completed: dict[str, dict] = cp.get("completed", {}) if cp else {}
    cp_total = cp.get("total", len(to_gen)) if cp else len(to_gen)
    previously_success = {u for u, info in cp_completed.items() if info.get("result") == "success"}
    previously_failed = {u for u, info in cp_completed.items() if info.get("result") == "failed"}
    to_gen = [u for u in to_gen if u not in previously_success]
    if cp_completed:
        print(f"   checkpoint resume: {len(previously_success)} skipped (OK), "
              f"{len(previously_failed)} skipped (failed), {len(to_gen)} remaining", flush=True)

    gen_results: dict[str, str | None] = {}
    _gen_lock = threading.Lock()
    _cp_done = 0

    # Multi-key gen pipeline with retry
    def gen_one(url: str) -> tuple[str, str | None]:
        # edits → nano-banana-2 → Doubao → GPT generations
        r = _try_key_chain(lambda: edit_gen(url, rr_edits.next_key() or hfsy_key))
        if r:
            return url, r
        r = _try_key_chain(lambda: nano_gen(url, rr_nano.next_key() or hfsy_key, size=gen_size))
        if r:
            return url, r
        r = _try_key_chain(lambda: doubao_gen(url, rr_doubao.next_key() or ark_key))
        if r:
            return url, r
        r = _try_key_chain(lambda: hfsyapi_gen(url, rr_hfsyapi.next_key() or hfsy_key))
        if r:
            return url, r
        return url, None

    with ThreadPoolExecutor(max_workers=gen_workers) as ex:
        fut = {ex.submit(gen_one, url): url for url in to_gen}
        _loop_start = time.monotonic()
        for f in as_completed(fut):
            orig_url, new_url = f.result()
            with _gen_lock:
                _cp_done += 1
                gen_results[orig_url] = new_url
                if _cp_done % 5 == 0:
                    _save_cp = dict(cp_completed)
                    _save_cp[orig_url] = {
                        "result": "success" if new_url else "failed",
                        "new_url": new_url or "",
                        "finished_at": time.time(),
                    }
                    checkpoint_save(ckpt_path, _save_cp, cp_total)
                elapsed = time.monotonic() - _loop_start
                _print_progress(_cp_done, len(to_gen), elapsed, "gen")

    # Global retry pool: retry failed images up to 2 more times with different keys
    failed_urls = [u for u, r in gen_results.items() if not r]
    max_retries = 2
    retry_round = 0
    while failed_urls and retry_round < max_retries:
        retry_round += 1
        # Reset key RRs for fresh key rotation
        rr_edits = KeyRoundRobin(hfsy_keys) if hfsy_keys else None
        rr_nano = KeyRoundRobin(hfsy_keys) if hfsy_keys else None
        rr_doubao = KeyRoundRobin(ark_keys) if ark_keys else None
        rr_hfsyapi = KeyRoundRobin(hfsy_keys) if hfsy_keys else None
        print(f"   Retry round {retry_round}/{max_retries}: {len(failed_urls)} failed images...", flush=True)
        time.sleep(3)  # brief pause before retry

        with ThreadPoolExecutor(max_workers=min(gen_workers, len(failed_urls))) as ex:
            fut = {ex.submit(gen_one, url): url for url in failed_urls}
            _retry_start = time.monotonic()
            _retry_done = 0
            for f in as_completed(fut):
                orig_url, new_url = f.result()
                with _gen_lock:
                    gen_results[orig_url] = new_url
                    _retry_done += 1
                    elapsed = time.monotonic() - _retry_start
                    _print_progress(_retry_done, len(failed_urls), elapsed, "retry")
                    if new_url:
                        # Success in retry — update checkpoint
                        _save_cp = dict(cp_completed)
                        _save_cp[orig_url] = {
                            "result": "success",
                            "new_url": new_url,
                            "finished_at": time.time(),
                            "retry_round": retry_round,
                        }
                        checkpoint_save(ckpt_path, _save_cp, cp_total)

        failed_urls = [u for u, r in gen_results.items() if not r]

    gen_ok = sum(1 for v in gen_results.values() if v)
    gen_fail = sum(1 for v in gen_results.values() if not v)
    print(f"   generated {gen_ok} OK, {gen_fail} failed (kept original)", flush=True)

    # Clean up checkpoint after successful completion
    if Path(ckpt_path).exists():
        Path(ckpt_path).unlink()

    # Apply results to work.json
    for row in w["rows"]:
        for img in row["images"]:
            url = img["orig"]
            src = img.get("source", "desc")
            d = decisions.get(url, "keep")
            if d == "delete" and src != "desc":
                d = "regen" if gen_results.get(url) else "keep"
            if d == "delete":
                img["decision"] = "delete"
            elif d == "regen":
                new_url = gen_results.get(url)
                if new_url:
                    img["decision"] = "regen"
                    img["new_url"] = new_url
                else:
                    img["decision"] = "keep"
                    img["new_url"] = ""
            else:
                img["decision"] = "keep"
                img["new_url"] = ""

    work.write_text(json.dumps(w, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"   saved {work}", flush=True)

    # Step 5: Finalize
    print("[4/5] Finalize (writing sheet)...", flush=True)
    rp_finalize(str(xlsx), str(work), str(xlsx))
    print("[5/5] Done!", flush=True)

    return {
        "rows": len(w["rows"]),
        "unique_images": len(all_urls),
        "delete": n_del, "regen": n_regen, "keep": n_keep,
        "generated_ok": gen_ok, "generated_fail": gen_fail,
    }


def _parse_keys(key_str: str) -> list[str]:
    """Parse comma-separated key string. Returns list of non-blank keys."""
    if not key_str:
        return []
    return [k.strip() for k in key_str.split(",") if k.strip()]


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser(description="Auto batch process TikTok Vietnam product sheet")
    ap.add_argument("xlsx", help="Path to xlsx file")
    ap.add_argument("--doubao-key", default=os.environ.get("ARK_API_KEY", ""),
                    help="Doubao/minimax-m3 key(s), comma-separated for rotation")
    ap.add_argument("--hfsy-key", default=os.environ.get("HFSY_API_KEY", ""),
                    help="HFSy API key(s), comma-separated for rotation")
    ap.add_argument("--agnes-key", default=os.environ.get("AGNES_API_KEY", ""),
                    help="Agnes vision key(s), comma-separated for rotation")
    ap.add_argument("--work", default=None)
    ap.add_argument("--audit-workers", type=int, default=24, help="parallel vision audits")
    ap.add_argument("--gen-workers", type=int, default=0,
                    help="parallel image generations (0=default to min(50, CPU*2))")
    ap.add_argument("--gen-size", default="4K", choices=["1K", "2K", "4K"])
    ap.add_argument("--checkpoint", default=None,
                    help="Checkpoint file path for resume support")
    args = ap.parse_args(argv)

    # Parse multi-key support
    ark_keys = _parse_keys(args.doubao_key)
    hfsy_keys = _parse_keys(args.hfsy_key)
    agnes_keys = _parse_keys(args.agnes_key)

    # Compute default workers
    if args.gen_workers <= 0:
        try:
            ncpu = len(os.sched_getaffinity(0))
        except (AttributeError, OSError):
            ncpu = os.cpu_count() or 1
        args.gen_workers = min(50, ncpu * 2)

    print("=== TK-VN Product Sheet Auto Processor ===", flush=True)
    print(f"Input: {args.xlsx}", flush=True)
    print(f"HFSy keys: {len(hfsy_keys)}  Doubao keys: {len(ark_keys)}  Agnes keys: {len(agnes_keys)}", flush=True)
    print(f"nano-banana-2/GPT-Image-2: {'ok' if hfsy_keys else 'MISSING'}  "
          f"Doubao: {'ok' if ark_keys else 'MISSING'}  "
          f"Vision: {'ok' if agnes_keys else 'MISSING'}", flush=True)
    print(f"Concurrency: audit={args.audit_workers} gen={args.gen_workers} size={args.gen_size}\n", flush=True)

    report = auto_process(
        args.xlsx,
        args.doubao_key, args.hfsy_key, args.agnes_key,
        args.work, args.audit_workers, args.gen_workers, args.gen_size,
        _ark_keys=ark_keys, _hfsy_keys=hfsy_keys,
        _checkpoint_path=args.checkpoint,
    )
    print("\n=== Summary ===", flush=True)
    for k, v in report.items():
        print(f"  {k}: {v}", flush=True)
    return 0
  
if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
