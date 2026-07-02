"""Sheet-level checks used by the eval runner.

Usage: python check_sheet.py <xlsx> <check_name>

check_name ∈ {brand_set, stock_set, video_cleared, sku_format, image_urls_https}
Exit 0 if the check passes, 1 otherwise. Prints a short reason on failure.
"""
from __future__ import annotations

import re
import sys

import openpyxl

SHEET = "tiktok_chanpin_"


def data_rows(ws) -> list[int]:
    return [
        r for r in range(2, ws.max_row + 1)
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, ws.max_column + 1))
    ]


def check_brand_set(ws) -> bool:
    # 品牌列留空(不填任何值)
    return all(not ws.cell(r, 4).value for r in data_rows(ws))


def check_stock_set(ws) -> bool:
    return all(ws.cell(r, 17).value == 30 for r in data_rows(ws))


def check_video_cleared(ws) -> bool:
    return all(not ws.cell(r, 27).value for r in data_rows(ws))


def check_sku_format(ws) -> bool:
    vals = [ws.cell(r, 6).value for r in data_rows(ws)]
    if not all(isinstance(v, str) and re.fullmatch(r"\d{13}", v) for v in vals):
        return False
    return vals == sorted(vals)


def check_image_urls_https(ws) -> bool:
    img_cols = [18] + list(range(19, 27)) + [29]
    for r in data_rows(ws):
        for c in img_cols:
            v = ws.cell(r, c).value
            if v and not str(v).startswith("https://"):
                return False
        desc = ws.cell(r, 3).value or ""
        for u in re.findall(r'<img[^>]+src="([^"]+)"', desc):
            if not u.startswith("https://"):
                return False
    return True


CHECKS = {
    "brand_set": check_brand_set,
    "stock_set": check_stock_set,
    "video_cleared": check_video_cleared,
    "sku_format": check_sku_format,
    "image_urls_https": check_image_urls_https,
}


def main(argv: list[str]) -> int:
    if len(argv) != 3 or argv[2] not in CHECKS:
        print(f"usage: check_sheet.py <xlsx> <{'|'.join(CHECKS)}>")
        return 2
    wb = openpyxl.load_workbook(argv[1], data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    ok = CHECKS[argv[2]](ws)
    print(f"{argv[2]}: {'PASS' if ok else 'FAIL'}")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main(sys.argv))
